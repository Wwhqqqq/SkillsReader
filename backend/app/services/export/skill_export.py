"""Export skills to CSV / Excel."""

from __future__ import annotations

import csv
import io
import zipfile
from datetime import date, datetime, timedelta, timezone
from typing import Any

from sqlalchemy import distinct, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Skill
from app.core.redis_client import get_official_scan_state
from app.services.enrichment.skill_classification import (
    data_source_for,
    publisher_type_for,
)
from app.services.scan.skill_gate import is_real_skill

DOMESTIC_VENDORS = ("美团", "阿里", "字节", "知乎", "小红书", "哔哩哔哩", "快手", "滴滴", "拼多多", "携程", "得物", "腾讯")
EXCLUDED_EXPORT_VENDORS: frozenset[str] = frozenset()
RECENT_EXPORT_EXCLUDE_VENDORS = frozenset({"海外社区", "GitHub"})
SKILLS_SH_SOURCE_ID = "skills_sh"
# 与精选 🆕 一致：first_seen 在 1 天内视为新发现
RECENT_EXPORT_HOURS = 24

EXPORT_COLUMNS_BASE: list[tuple[str, str]] = [
    ("序号", "index"),
    ("Skill名称", "name"),
    ("公司", "vendor"),
    ("发布类型", "publisher_type"),
    ("数据来源", "data_source"),
    ("分类", "category"),
    ("描述", "description"),
    ("链接", "detail_url"),
    ("发布日期", "publish_date"),
    ("首次发现(北京)", "first_seen_at"),
    ("来源", "source_id"),
]

EXPORT_METRIC_COLUMNS: list[tuple[str, str]] = [
    ("安装量", "install_count"),
    ("质量分", "quality_score"),
]

XLSX_WIDTHS_BASE = [8, 28, 10, 12, 14, 14, 48, 36, 12, 20, 18]
XLSX_WIDTHS_METRICS = [10, 10]


def _shanghai_tz():
    from zoneinfo import ZoneInfo

    return ZoneInfo("Asia/Shanghai")


def _shanghai_now() -> datetime:
    try:
        return datetime.now(_shanghai_tz())
    except Exception:
        return datetime.now().astimezone(timezone.utc)


def _today_shanghai() -> date:
    return _shanghai_now().date()


def _recent_hours_cutoff_utc(hours: int, *, now_shanghai: datetime | None = None) -> datetime:
    """上海时区往前 hours 小时，转为与 first_seen_at 一致的 UTC naive。"""
    ref_sh = now_shanghai or _shanghai_now()
    cutoff_sh = ref_sh - timedelta(hours=hours)
    return cutoff_sh.astimezone(timezone.utc).replace(tzinfo=None)


def _recent_hours_filter(hours: int):
    return Skill.first_seen_at >= _recent_hours_cutoff_utc(hours)


def _today_ingest_start_utc() -> datetime:
    """上海时区当日 0:00 对应的 UTC naive 时间（与 first_seen_at 存储一致）。"""
    now_sh = _shanghai_now()
    start_sh = now_sh.replace(hour=0, minute=0, second=0, microsecond=0)
    return start_sh.astimezone(timezone.utc).replace(tzinfo=None)


def _today_ingest_filter():
    """当日 0:00（上海）至今入库：按 first_seen_at。"""
    return Skill.first_seen_at >= _today_ingest_start_utc()


def _is_exportable(skill: Skill) -> bool:
    return is_real_skill(skill)


def export_columns_for(skills: list[Skill]) -> list[tuple[str, str]]:
    cols = list(EXPORT_COLUMNS_BASE)
    if _is_skills_sh_export(skills):
        cols = cols[:-1] + EXPORT_METRIC_COLUMNS + [cols[-1]]
    return cols


def _is_skills_sh_export(skills: list[Skill]) -> bool:
    return bool(skills) and all(s.source_id == SKILLS_SH_SOURCE_ID for s in skills)


def _skill_category(skill: Skill) -> str:
    meta = skill.metadata_json or {}
    if isinstance(meta, dict):
        if meta.get("categoryName"):
            return str(meta["categoryName"])
        if meta.get("category"):
            return str(meta["category"])
    tags = skill.tags or []
    skip = {"美团", "阿里", "腾讯", "字节", "知乎", "小红书", "哔哩哔哩", "快手", "滴滴", "拼多多", "携程", "得物", "github", "skills.sh", "云资源", "生活服务"}
    for tag in tags:
        if tag not in skip:
            return str(tag)
    return str(tags[0]) if tags else "-"


def _format_first_seen_at(dt: datetime | None) -> str:
    """first_seen_at 存 UTC naive，导出为北京时间。"""
    if not dt:
        return ""
    utc = dt.replace(tzinfo=timezone.utc)
    return utc.astimezone(_shanghai_tz()).strftime("%Y-%m-%d %H:%M:%S")


def skill_to_row(skill: Skill, index: int) -> dict[str, Any]:
    desc = (skill.llm_summary or skill.raw_description or "").strip()
    first_seen = _format_first_seen_at(skill.first_seen_at)
    publish = skill.publish_date.isoformat() if skill.publish_date else ""
    return {
        "index": index,
        "name": skill.name,
        "vendor": skill.vendor,
        "publisher_type": publisher_type_for(skill),
        "data_source": data_source_for(skill),
        "category": _skill_category(skill),
        "description": desc,
        "detail_url": skill.detail_url or "",
        "publish_date": publish,
        "first_seen_at": first_seen,
        "install_count": skill.install_count or 0,
        "quality_score": skill.quality_score or 0,
        "source_id": skill.source_id,
    }


def skill_in_recent_hours_window(
    skill: Skill,
    *,
    hours: int,
    now_shanghai: datetime | None = None,
) -> bool:
    if skill.first_seen_at is None:
        return False
    cutoff = _recent_hours_cutoff_utc(hours, now_shanghai=now_shanghai)
    return skill.first_seen_at >= cutoff


async def fetch_skills_for_export(
    session: AsyncSession,
    *,
    vendor: str | None = None,
    today_only: bool = False,
    recent_only: bool = False,
    last_official_scan: bool = False,
) -> list[Skill]:
    if last_official_scan:
        return await fetch_skills_for_last_official_scan(session, vendor=vendor)
    if vendor and vendor in EXCLUDED_EXPORT_VENDORS:
        return []

    query = select(Skill).where(Skill.status == "active")
    query = query.where(Skill.vendor.notin_(EXCLUDED_EXPORT_VENDORS))
    if vendor:
        query = query.where(Skill.vendor == vendor)
    elif recent_only:
        query = query.where(Skill.vendor.notin_(RECENT_EXPORT_EXCLUDE_VENDORS))
    if recent_only:
        query = query.where(_recent_hours_filter(RECENT_EXPORT_HOURS))
    elif today_only:
        query = query.where(_today_ingest_filter())
    skills = list(
        (
            await session.scalars(query.order_by(Skill.first_seen_at.desc()))
        ).all()
    )
    return [s for s in skills if _is_exportable(s)]


async def fetch_skills_for_last_official_scan(
    session: AsyncSession,
    *,
    vendor: str | None = None,
) -> list[Skill]:
    """导出最近一次官方一键扫描新增的官方 Skill。"""
    state = await get_official_scan_state()
    if state.get("status") not in ("done", "error"):
        return []
    ids = state.get("new_official_skill_ids") or state.get("new_skill_ids") or []
    if not ids:
        return []
    query = select(Skill).where(Skill.id.in_(ids), Skill.status == "active")
    if vendor:
        query = query.where(Skill.vendor == vendor)
    skills = list((await session.scalars(query.order_by(Skill.first_seen_at.desc()))).all())
    return [s for s in skills if _is_exportable(s)]


def rows_to_csv(skills: list[Skill]) -> bytes:
    columns = export_columns_for(skills)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for label, _ in columns])
    for i, skill in enumerate(skills, 1):
        row = skill_to_row(skill, i)
        writer.writerow([row[key] for _, key in columns])
    return "\ufeff".encode() + buf.getvalue().encode("utf-8")


def rows_to_xlsx(skills: list[Skill]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    columns = export_columns_for(skills)
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append([label for label, _ in columns])
    for i, skill in enumerate(skills, 1):
        row = skill_to_row(skill, i)
        ws.append([row[key] for _, key in columns])

    widths = list(XLSX_WIDTHS_BASE)
    if _is_skills_sh_export(skills):
        widths = widths[:-1] + XLSX_WIDTHS_METRICS + [widths[-1]]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_export_filename(
    *,
    fmt: str,
    vendor: str | None,
    today_only: bool,
    recent_only: bool = False,
    last_official_scan: bool = False,
) -> str:
    scope = vendor or "全部公司"
    if last_official_scan:
        kind = "最近官方扫描新增"
    elif recent_only:
        kind = f"近{RECENT_EXPORT_HOURS}小时新发现"
    elif today_only:
        kind = "今日新增"
    else:
        kind = "全量"
    ext = "xlsx" if fmt == "xlsx" else "csv"
    safe_vendor = scope.replace("/", "-")
    date_str = _today_shanghai().isoformat()
    return f"SkillGetter_{safe_vendor}_{kind}_{date_str}.{ext}"


def build_bundle_filename(
    *,
    today_only: bool,
    scope: str = "domestic",
    recent_only: bool = False,
) -> str:
    if recent_only:
        kind = f"近{RECENT_EXPORT_HOURS}小时新发现"
    elif today_only:
        kind = "今日新增"
    else:
        kind = "全量"
    scope_label = "国内公司" if scope == "domestic" else "全部公司"
    date_str = _today_shanghai().isoformat()
    return f"SkillGetter_{scope_label}_{kind}_{date_str}.zip"


async def list_export_vendors(
    session: AsyncSession,
    *,
    scope: str = "domestic",
) -> list[str]:
    rows = await session.scalars(
        select(distinct(Skill.vendor))
        .where(Skill.status == "active")
        .where(Skill.vendor.notin_(EXCLUDED_EXPORT_VENDORS))
        .order_by(Skill.vendor)
    )
    vendors = [v for v in rows.all() if v]
    if scope == "domestic":
        preferred = [v for v in DOMESTIC_VENDORS if v in vendors]
        extra = [
            v
            for v in vendors
            if v not in DOMESTIC_VENDORS and v not in ("海外社区", "GitHub")
        ]
        return preferred + extra
    return vendors


async def build_vendor_bundle(
    session: AsyncSession,
    *,
    today_only: bool = False,
    recent_only: bool = False,
    scope: str = "domestic",
    vendors: list[str] | None = None,
    fmt: str = "csv",
) -> tuple[bytes, list[str]]:
    """One file per company, packaged as ZIP. Returns (zip_bytes, company names)."""
    all_vendors = await list_export_vendors(session, scope=scope)
    if vendors:
        wanted = set(vendors)
        vendor_list = [v for v in all_vendors if v in wanted]
        extra = [v for v in vendors if v not in vendor_list]
        vendor_list.extend(extra)
    else:
        vendor_list = all_vendors

    buf = io.BytesIO()
    included: list[str] = []
    ext = "xlsx" if fmt == "xlsx" else "csv"

    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for vendor in vendor_list:
            skills = await fetch_skills_for_export(
                session,
                vendor=vendor,
                today_only=today_only,
                recent_only=recent_only,
            )
            if not skills:
                continue
            filename = build_export_filename(
                fmt=fmt,
                vendor=vendor,
                today_only=today_only,
                recent_only=recent_only,
            )
            content = rows_to_xlsx(skills) if fmt == "xlsx" else rows_to_csv(skills)
            zf.writestr(filename, content)
            included.append(vendor)

    return buf.getvalue(), included


async def build_vendor_csv_bundle(
    session: AsyncSession,
    *,
    today_only: bool = False,
    recent_only: bool = False,
    scope: str = "domestic",
    vendors: list[str] | None = None,
) -> tuple[bytes, list[str]]:
    """One CSV per company, packaged as ZIP. Returns (zip_bytes, company names)."""
    return await build_vendor_bundle(
        session,
        today_only=today_only,
        recent_only=recent_only,
        scope=scope,
        vendors=vendors,
        fmt="csv",
    )
